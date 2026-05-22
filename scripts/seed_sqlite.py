import os
import sqlite3
import uuid
import openpyxl
from datetime import datetime
import re

DB_PATH = "warehouse.db"
EXCEL_PATH = "URE.xlsx"

# Helper to normalize strings for ID generation
def generate_slug(text):
    if not text:
        return str(uuid.uuid4())
    # Simple accent removal and lowercasing for Vietnamese
    text = text.lower().strip()
    replacements = {
        'à':'a','á':'a','ả':'a','ã':'a','ạ':'a','ă':'a','ằ':'a','ắ':'a','ẳ':'a','ẵ':'a','ặ':'a','â':'a','ầ':'a','ấ':'a','ẩ':'a','ẫ':'a','ậ':'a',
        'đ':'d',
        'è':'e','é':'e','ẻ':'e','ẽ':'e','ẹ':'e','ê':'e','ề':'e','ế':'e','ể':'e','ễ':'e','ệ':'e',
        'ì':'i','í':'i','ỉ':'i','ĩ':'i','ị':'i',
        'ò':'o','ó':'o','ỏ':'o','õ':'o','ọ':'o','ô':'o','ồ':'o','ố':'o','ổ':'o','ỗ':'o','ộ':'o','ơ':'o','ờ':'o','ớ':'o','ở':'o','ỡ':'o','ợ':'o',
        'ù':'u','ú':'u','ủ':'u','ũ':'u','ụ':'u','ư':'u','ừ':'u','ứ':'u','ử':'u','ữ':'u','ự':'u',
        'ỳ':'y','ý':'y','ỷ':'y','ỹ':'y','ỵ':'y'
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    text = re.sub(r'[^a-z0-9\s-]', '', text)
    text = re.sub(r'[\s-]+', '-', text)
    return text.strip('-')

def main():
    print("--- Starting SQLite Database Seeding ---")
    
    # 1. Connect to SQLite
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 2. Create Tables
    cursor.execute("DROP TABLE IF EXISTS transactions;")
    cursor.execute("DROP TABLE IF EXISTS drivers;")
    cursor.execute("DROP TABLE IF EXISTS materials;")
    
    cursor.execute("""
    CREATE TABLE drivers (
        id TEXT PRIMARY KEY,
        code TEXT UNIQUE NOT NULL,
        plate TEXT,
        name TEXT
    );
    """)
    
    cursor.execute("""
    CREATE TABLE materials (
        id TEXT PRIMARY KEY,
        code TEXT UNIQUE,
        name TEXT NOT NULL,
        unit TEXT,
        location TEXT,
        initial_quantity REAL DEFAULT 0,
        check_month TEXT
    );
    """)
    
    cursor.execute("""
    CREATE TABLE transactions (
        id TEXT PRIMARY KEY,
        date TEXT NOT NULL,
        type TEXT NOT NULL, -- 'NHAP' or 'XUAT'
        material_id TEXT NOT NULL,
        quantity REAL NOT NULL,
        driver_id TEXT, -- nullable
        odometer REAL, -- nullable
        notes TEXT,
        FOREIGN KEY (material_id) REFERENCES materials (id),
        FOREIGN KEY (driver_id) REFERENCES drivers (id)
    );
    """)
    
    # 3. Load Excel Workbook
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: {EXCEL_PATH} not found!")
        return
        
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # 4. Extract Drivers (from Trang tính2 lookup table K5:M17)
    s2 = wb['Trang tính2']
    drivers_data = []
    # Row 4 is header, row 5 to 17 are driver records
    for row_idx in range(5, 18):
        code = s2.cell(row=row_idx, column=11).value # Col K (11)
        plate = s2.cell(row=row_idx, column=12).value # Col L (12)
        name = s2.cell(row=row_idx, column=13).value # Col M (13)
        if code or plate or name:
            driver_id = f"drv-{generate_slug(code)}"
            drivers_data.append((driver_id, code, plate, name))
            
    # Add an extra fallback driver for security/guards (Để B.Vệ / Dp) if not already added
    if not any(d[1] == 'Dp' for d in drivers_data):
        drivers_data.append(("drv-dp", "Dp", "Dự phòng", "Để B.Vệ"))
        
    print(f"Extracted {len(drivers_data)} drivers.")
    cursor.executemany("INSERT INTO drivers (id, code, plate, name) VALUES (?, ?, ?, ?);", drivers_data)
    
    # Create driver map for fast lookup
    driver_map = {d[1].lower(): d[0] for d in drivers_data}
    # Also add driver by plate lookup
    driver_plate_map = {d[2].replace(" ", "").lower(): d[0] for d in drivers_data if d[2]}
    
    # 5. Extract Materials (from Trang tính1 and Trang tính2)
    s1 = wb['Trang tính1']
    raw_materials = set()
    
    # Scan Trang tính1 for unique materials
    for row in list(s1.iter_rows(values_only=True))[1:]:
        mat_name = row[1]
        if mat_name:
            raw_materials.add(mat_name.strip())
            
    print(f"Found {len(raw_materials)} raw material names in sheet 1.")
    
    materials_data = []
    # Explicitly define Ure (Lít) and Ure (Bình)
    # We will map "Ure" in sheet 1 (which is jerry cans/bottles) to "Ure (Bình)"
    # We will map Ure in sheet 2 (which is liters) to "Ure (Lít)"
    ure_lit_id = "mat-ure-lit"
    ure_binh_id = "mat-ure-binh"
    
    materials_data.append((ure_lit_id, "URE-L", "Ure (Lít)", "Lít", "Bồn chứa", 0.0, "2026-04"))
    materials_data.append((ure_binh_id, "URE-B", "Ure (Bình)", "Bình", "Kho phụ tùng", 0.0, "2025-11"))
    
    material_map = {
        "ure": ure_binh_id # Map exact 'Ure' from sheet 1 to 'Ure (Bình)'
    }
    
    used_codes = {"URE-L", "URE-B"}
    used_ids = {ure_lit_id, ure_binh_id}
    
    for mat in sorted(raw_materials):
        if mat.lower() == 'ure':
            continue
        slug = generate_slug(mat)
        mat_id = f"mat-{slug}"
        counter_id = 1
        base_id = mat_id
        while mat_id in used_ids:
            mat_id = f"{base_id}-{counter_id}"
            counter_id += 1
            
        used_ids.add(mat_id)
        
        # Determine unit from sheet 1
        unit = "Cái" # default
        for row in list(s1.iter_rows(values_only=True))[1:]:
            if row[1] and row[1].strip() == mat:
                unit = row[4] or "Cái"
                break
        
        # Simple rule to estimate location
        location = "Kho chung"
        if "sơn" in mat.lower() or "silicon" in mat.lower():
            location = "Tủ hóa chất"
        elif "đinh" in mat.lower() or "cảo" in mat.lower() or "rotin" in mat.lower() or "bóng" in mat.lower():
            location = "Tủ phụ tùng"
            
        base_code = f"VT-{slug[:5].upper()}"
        code = base_code
        counter = 1
        while code in used_codes:
            suffix = f"-{counter}"
            code = f"VT-{slug[:5].upper()}{suffix}"
            counter += 1
            
        used_codes.add(code)
        materials_data.append((mat_id, code, mat, unit, location, 0.0, "2025-11"))
        material_map[mat.lower()] = mat_id
        
    print(f"Prepared {len(materials_data)} materials for insertion.")
    cursor.executemany("INSERT INTO materials (id, code, name, unit, location, initial_quantity, check_month) VALUES (?, ?, ?, ?, ?, ?, ?);", materials_data)
    
    # 6. Extract Transactions from Trang tính1 (General Warehouse)
    transactions_data = []
    # Sheet 1: [Ngày, Tên Vật Tư, SL Nhập kho, SL Xuất kho, Đơn vị, Ghi chú]
    s1_rows = list(s1.iter_rows(values_only=True))[1:]
    
    for idx, row in enumerate(s1_rows):
        dt_val = row[0]
        mat_name = row[1]
        sl_nhap = row[2]
        sl_xuat = row[3]
        unit = row[4]
        ghi_chu = row[5] or ''
        
        if not dt_val or not mat_name:
            continue
            
        # Parse date
        if isinstance(dt_val, datetime):
            date_str = dt_val.strftime("%Y-%m-%d")
        else:
            date_str = str(dt_val)[:10]
            
        # Map material
        mat_id = material_map.get(mat_name.strip().lower())
        if not mat_id:
            continue
            
        tx_id = f"tx-s1-{idx}-{uuid.uuid4().hex[:6]}"
        tx_type = 'NHAP' if sl_nhap is not None and float(sl_nhap) > 0 else 'XUAT'
        qty = float(sl_nhap) if tx_type == 'NHAP' else float(sl_xuat)
        
        # Try to map driver/vehicle from note
        drv_id = None
        # Look for driver code or vehicle plate in note
        note_lower = ghi_chu.lower().replace(" ", "")
        
        # 1. Look for license plates like 50H25543, 50H93044
        found_plate = None
        plate_match = re.search(r'\d{2}[a-z\d]\d{4,5}', note_lower)
        if plate_match:
            found_plate = plate_match.group(0)
            
        # 2. Look for abbreviated plates like 93044, 25543, 06134
        if not found_plate:
            num_match = re.search(r'\d{5}', note_lower)
            if num_match:
                # Find matching driver plate that ends with this number
                end_num = num_match.group(0)
                for drv in drivers_data:
                    drv_plate_clean = drv[2].replace(" ", "").lower() if drv[2] else ""
                    if drv_plate_clean.endswith(end_num):
                        drv_id = drv[0]
                        break
        
        if found_plate and not drv_id:
            # Map plate to driver
            for drv in drivers_data:
                drv_plate_clean = drv[2].replace(" ", "").lower() if drv[2] else ""
                if found_plate in drv_plate_clean or drv_plate_clean in found_plate:
                    drv_id = drv[0]
                    break
                    
        # 3. Direct driver code or name match
        if not drv_id:
            for drv in drivers_data:
                drv_code_clean = drv[1].lower()
                drv_name_clean = drv[3].lower().replace(" ", "")
                if drv_code_clean in note_lower or drv_name_clean in note_lower:
                    drv_id = drv[0]
                    break
                    
        transactions_data.append((tx_id, date_str, tx_type, mat_id, qty, drv_id, None, ghi_chu))
        
    print(f"Extracted {len(transactions_data)} transactions from sheet 1.")
    
    # 7. Extract Transactions from Trang tính2 (Ure Specialized)
    s2_rows = list(s2.iter_rows(values_only=True))[1:]
    s2_tx_count = 0
    
    for idx, row in enumerate(s2_rows):
        dt_val = row[0]
        tx_code = row[1]
        xuat_lit = row[3]
        nhap_lit = row[4]
        so_km = row[5]
        
        if dt_val is None or (tx_code is None and xuat_lit is None and nhap_lit is None):
            continue
            
        # Parse date
        if isinstance(dt_val, datetime):
            date_str = dt_val.strftime("%Y-%m-%d")
        else:
            date_str = str(dt_val)[:10]
            
        tx_id = f"tx-s2-{idx}-{uuid.uuid4().hex[:6]}"
        
        if nhap_lit is not None:
            # Refill Ure in bulk
            tx_type = 'NHAP'
            qty = float(nhap_lit)
            drv_id = None
            odometer = None
            notes = "Nhập bồn chứa Ure bồn lớn"
        else:
            # Export Ure to vehicle
            tx_type = 'XUAT'
            qty = float(xuat_lit) if xuat_lit is not None else 0.0
            drv_id = driver_map.get(str(tx_code).strip().lower()) if tx_code else None
            odometer = float(so_km) if so_km is not None else None
            notes = f"Đổ Ure xe {row[2] or ''} - Tài xế: {row[6] or ''}"
            
        transactions_data.append((tx_id, date_str, tx_type, ure_lit_id, qty, drv_id, odometer, notes))
        s2_tx_count += 1
        
    print(f"Extracted {s2_tx_count} Ure transactions from sheet 2.")
    
    # 8. Insert All Transactions
    cursor.executemany("""
    INSERT INTO transactions (id, date, type, material_id, quantity, driver_id, odometer, notes)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?);
    """, transactions_data)
    
    # Commit and Close
    conn.commit()
    conn.close()
    
    print("\n✅ Seeding SQLite Database Successful!")
    print(f"Database created at: {os.path.abspath(DB_PATH)}")
    print(f"Total transactions seeded: {len(transactions_data)}")

if __name__ == "__main__":
    main()
